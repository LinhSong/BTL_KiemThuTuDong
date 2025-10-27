import openpyxl
import os
from datetime import datetime

class ExcelUtils:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)
        self.sheet = self.wb[sheet_name]

    def get_data(self, mode="login"):
        data = []
        for row in range(2, self.sheet.max_row + 1):
            if mode == "login":
                tc_id = self.sheet.cell(row, 1).value or ""
                email = self.sheet.cell(row, 2).value or ""
                password = self.sheet.cell(row, 3).value or ""
                expected = self.sheet.cell(row, 4).value or ""
                data.append((tc_id, email, password, expected))

            elif mode == "register":
                tc_id = self.sheet.cell(row, 1).value or ""
                ho = self.sheet.cell(row, 2).value or ""
                ten = self.sheet.cell(row, 3).value or ""
                email = self.sheet.cell(row, 4).value or ""
                sdt = self.sheet.cell(row, 5).value or ""
                password = str(self.sheet.cell(row, 6).value or "")
                expected = self.sheet.cell(row, 7).value or ""
                data.append((tc_id, ho, ten, email, sdt, password, expected))

            elif mode == "DoiMatKhau":
                tc_id = self.sheet.cell(row, 1).value or ""
                old_pass = str(self.sheet.cell(row, 2).value or "")
                new_pass = str(self.sheet.cell(row, 3).value or "")
                confirm_pass = str(self.sheet.cell(row, 4).value or "")
                expected = self.sheet.cell(row, 5).value or ""
                data.append((tc_id, old_pass, new_pass, confirm_pass, expected))
                
            elif mode == "search":
                tc_id = self.sheet.cell(row, 1).value or ""
                keyword = self.sheet.cell(row, 2).value or ""
                expected = self.sheet.cell(row, 3).value or ""
                data.append((tc_id, keyword, expected))
            elif mode == "cart":
                    tc_id = self.sheet.cell(row, 1).value or ""
                    keyword = self.sheet.cell(row, 2).value or ""
                    expected = self.sheet.cell(row, 3).value or ""
                    data.append((tc_id, keyword, expected))
        return data

    def write_result(self, tc_id, actual, expected, mode="login", screenshot_path=""):
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row, 1).value == tc_id:
                # So sánh và gán giá trị bool
                result = expected in actual
                if mode == "login":
                    self.sheet.cell(row, 5).value = actual               
                    self.sheet.cell(row, 6).value = "Pass" if expected in actual else "Fail"  
                    if screenshot_path:
                        self.sheet.cell(row, 7).value = screenshot_path
                elif mode == "register":
                    self.sheet.cell(row, 8).value = actual
                    self.sheet.cell(row, 9).value = "Pass" if expected in actual else "Fail"

                elif mode == "DoiMatKhau":
                    self.sheet.cell(row, 6).value = actual
                    self.sheet.cell(row, 7).value = "Pass" if expected in actual else "Fail"
                    if screenshot_path:  
                        self.sheet.cell(row, 8).value = screenshot_path
                elif mode == "search":
                    self.sheet.cell(row, 4).value = actual
                    self.sheet.cell(row, 5).value = "Pass" if expected in actual else "Fail"
                    if screenshot_path:
                        self.sheet.cell(row, 6).value = screenshot_path
                elif mode == "cart":
                    self.sheet.cell(row, 4).value = actual
                    self.sheet.cell(row, 5).value = "Pass" if expected in actual else "Fail"
                    if screenshot_path:
                        self.sheet.cell(row, 6).value = screenshot_path
    
                break
        self.wb.save(self.file_path)

    @staticmethod
    def screenshot(driver, tc_id):
        screenshots_dir = os.path.join(os.getcwd(), "screenshots")
        if not os.path.exists(screenshots_dir):
            os.makedirs(screenshots_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(screenshots_dir, f"{tc_id}_{timestamp}.png")
        driver.save_screenshot(file_path)
        return file_path
